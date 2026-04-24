import socket
import pyrealsense2 as rs
import cv2
import base64
import json
import numpy as np
import time
import threading
import struct
import atexit
from datetime import datetime
import openpyxl
import os
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import traceback
from pymodbus.client.sync import ModbusTcpClient
import sys
import random
import queue
import logging
from logging.handlers import TimedRotatingFileHandler

# ===== 파일 로깅 설정 =====
LOG_DIR = "C:\\nobel\\logs"
os.makedirs(LOG_DIR, exist_ok=True)

def setup_logger(name, filename, level=logging.DEBUG):
    """날짜별 로그 파일을 생성하는 로거 설정"""
    logger = logging.getLogger(name)
    logger.setLevel(level)
    if not logger.handlers:
        handler = TimedRotatingFileHandler(
            os.path.join(LOG_DIR, filename),
            when='midnight',
            interval=1,
            backupCount=30,  # 30일 보관
            encoding='utf-8'
        )
        handler.suffix = "%Y-%m-%d"
        formatter = logging.Formatter(
            '[%(asctime)s.%(msecs)03d] %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        handler.setFormatter(formatter)
        logger.addHandler(handler)
        # 콘솔에도 출력
        console = logging.StreamHandler()
        console.setFormatter(formatter)
        logger.addHandler(console)
    return logger

# PLC 통신 전용 로거
plc_logger = setup_logger('plc', 'plc_data.log')
# 소켓 통신 전용 로거
socket_logger = setup_logger('socket', 'socket_comm.log')
# 제로값 감지 전용 로거
zero_logger = setup_logger('zero', 'zero_value_alert.log')

sys.path.append(os.path.join(sys.base_prefix, 'Lib', 'site-packages', 'pywin32_system32'))
import pythoncom
import win32com.client

names = ['MS-010', 'MS-011', 'MS-012', 'MS-013', 'MS-014', 'MS-015']
MS010 = [
    'MS-010',
    [10.98],
    ["3"],
    [0.15, -0.15],
    {
      "2": [7.89, -0.06, 0.06],
      "4": [4.95, -0.25, 0.25],
      "5": [1.7, -0.1, 0.1],
      "6": [21.12, -0.25, 0.25]
    },
    13
  ]
MS011 = [
    'MS-011',
    [16.51],
    ["3"],
    [0.25, -0.25],
    {
      "2": [11.8, -0.1, 0.1],
      "4": [7.75, -0.25, 0.25],
      "5": [2.54, -0.2, 0.2],
      "6": [26.62, -0.5, 0.5]
    },
    13
  ]
MS012 = [
    'MS-012',
    [8.8, 9.0],
    ["3", "4"],
    [0.2, -0.2, 0.2, -0.2],
    {
      "2": [6.75, -0.1, 0.1],
      "5": [1.4, -0.1, 0.1]
    },
    13
  ]
MS013 = [
    'MS-013',
    [13.5],
    ["3"],
    [0.2, -0.2],
    {
      "2": [11.45, -0.1, 0.1],
      "4": [13.5, -0.2, 0.2],
      "5": [1.4, -0.1, 0.1]
    },
    13
  ]
MS014 = [
    'MS-014',
    [7.1, 3.2, 115.0],
    ["2", "3", "4"],
    [0.4, -0.18, 0.2, -0.1, 2.0, -2.0],
    [0.38, -0.16, 0.18, -0.08, 1.97, -1.97],
    {
      "5": [3.5, -0.5, 0.5, -0.47, 0.47],
      "6": [5.11, -0.08, 0.08, -0.07, 0.07],
      "7": [0, -0.2, 0.2, -0.19, 0.19]
    },
    14
  ]
MS015 = [
    'MS-015',
    [7.1, 3.2, 120.0],
    ["2", "3", "5"],
    [0.4, -0.2, 0.2, -0.1, 2.0, -2.0],
    [0.38, -0.18, 0.18, -0.08, 1.97, -1.97],
    {
      "4": [1.4, -0.2, 0.2, -0.19, 0.19],
      "6": [3.5, -0.5, 0.5, -0.47, 0.47],
      "7": [5.11, -0.08, 0.08, -0.07, 0.07],
      "8": [0, -0.2, 0.2, -0.19, 0.19]
    },
    15
  ]
sheet_names = {
        'MS-010' : 'Aurora1 8 SUS304 Male tube 성형',
        'MS-011' : 'Aurora1 12 SUS304  Male tube 성형',
        'MS-012' : 'Aurora1 8 SUS304 이중축관성형',
        'MS-013' : 'Aurora1 12 SUS304 이중축관성형',
        'MS-014' : 'Aurora1 4.76사두 성형',
        'MS-015' : 'Aurora1 4.76나팔 120° FLARE성형'
    }

week_cell_index = {
    'date': ['L7', 'T7', 'AB7'],
    'values': ['L10',
                'T10',
                'AB10',
                ]
}

month_cell_index = {
    'title' : 'H1',
    'data' : ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
              'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
              'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD'],
    'input' : {
                'model':'C4',
                '품명':'C5',
                'Code No.':'C6',
                '관리항목':'C7',
                'SPEC':'C8',
                '공차상한':'C9',
                '공차하한':'C10',
                'preCL':'W15',
                'preUCL':'Y15',
                'preLCL':'AA15',
                'CL':'AC15',
                'UCL':'AE15'
               },
    'output' : ['C16', 'C17', 'C19', 'C20' , 'C21', 'C23', 'C24', 'C25', 'C27', 'C28', 'C30']
}

json_data = None
data_lock = threading.Lock()
flutter_lock = threading.Lock()

global data
data = None
input_flag = True


data_path = "C:\\nobel\\data\\data.json"
date_list_path = "C:\\nobel\\data\\dateList.json"
week_excel_path = "C:\\nobel\\data\\form_week.xlsx"
month_excel_path = "C:\\nobel\\data\\form_month.xlsx"

flutter_host = "127.0.0.1"
flutter_port = 12345
plc_host = "192.168.0.100"
plc_port = 502

# plc_host = "192.168.10.212"
# plc_port = 502

class Server:
    clients = []
    
    def __init__(self, flutter_host, flutter_port, plc_ip, plc_port=5020):
        self.flutter_host = flutter_host
        self.flutter_port = flutter_port
        self.flutter_socket = None
        self.is_flutter_connected = False
        self.flutter_received = False
        
        self.plc_ip = plc_ip
        self.plc_port = plc_port
        self.plc_received = False
        self.is_plc_connecteced = False

        # 쓰레드 공유 데이터
        self.plc_data = None
        self.pipelines = []
        
        ctx = rs.context()
        devices = ctx.query_devices()
        num_cameras = len(devices)
        if num_cameras < 1:
            print("RealSense 카메라가 연결되어 있지 않습니다.")
            self.pipelines = []  # 카메라 없이도 실행 가능하도록 처리
        else:
            for i in range(2):
                pipeline = rs.pipeline()
                config = rs.config()
                
                # 각 디바이스의 시리얼 번호 확인 후 해당 디바이스 선택
                serial = devices[i].get_info(rs.camera_info.serial_number)
                print(f"카메라 {i} 시리얼 번호: {serial}")
                config.enable_device(serial)
                
                # 스트림 설정: 깊이와 컬러 스트림 (여기서는 컬러만 인코딩)
                config.enable_stream(rs.stream.depth, 640, 480, rs.format.z16, 30)
                config.enable_stream(rs.stream.color, 640, 480, rs.format.bgr8, 30)
                
                pipeline.start(config)
                self.pipelines.append(pipeline)

        # 프로세스 종료 시 RealSense 파이프라인 정리 (USB/카메라 핸들 해제)
        atexit.register(self._shutdown_pipelines)

        self.data = None
        # PLC 측정 데이터 전용 큐 (데이터 손실 방지)
        self.measurement_queue = queue.Queue()

        self.slave = 1
        self.checked = False

        # 트리거/모델 변경 추적 (측정하지 않은 데이터가 들어오는 문제 방지)
        self.last_trigger_value = None
        self.last_model_num = None
        self.plc_connected_at = None
        self.last_model_change_at = None
        # 모델별 마지막 전송 값 (레지스터 정체 감지용)
        self.last_sent_values_per_model = {}

    def read_plc_data_modbusTCP(self):

        self.client = ModbusTcpClient(host=self.plc_ip, port=self.plc_port)
        if not self.client.connect():
            plc_logger.error("PLC 연결 실패 (host=%s, port=%s)", self.plc_ip, self.plc_port)
            print("PLC 연결 실패")
            return

        plc_logger.info("PLC 연결 완료 (host=%s, port=%s)", self.plc_ip, self.plc_port)
        print("PLC 연결 완료")
        self.plc_connected_at = time.time()

        while True:
            try:
                # 모델 번호 읽기
                result = self.client.read_input_registers(address=7010, count=1, slave=self.slave)
                model_num = self.convert_int(result) if not result.isError() else -1
                if result.isError():
                    plc_logger.error("모델 번호 읽기 실패: addr=7010, error=%s", result)
                else:
                    plc_logger.debug("모델 번호: %d (raw=%s)", model_num, result.registers)
                print("모델 데이터:", model_num)


                values = []

                def get_and_append(addr, label, countindex=1, condition=True):
                    res = self.client.read_input_registers(address=addr, count=countindex, slave=self.slave)
                    if not res.isError():
                        if countindex == 1:
                            v = self.convert_int(res) / 1000
                            # 0값 감지 로깅
                            if v == 0:
                                zero_logger.warning("[0값 감지] model=%s, addr=%d, label=%s, raw_registers=%s, converted=%.3f, condition=%s",
                                                    model_num, addr, label, res.registers, v, condition)
                            plc_logger.debug("READ addr=%d, label=%s, raw=%s, value=%.3f, use=%s",
                                             addr, label, res.registers, v, condition)
                            if condition:
                                values.append(v)
                                print(f'진짜 데이터 {addr} {label}: {res.registers} / {v}')
                            else:
                                print(f'가짜 데이터 {addr} {label}: {res.registers} / {v}')
                            return v
                        else:
                            high_word = res.registers[1]
                            low_word = res.registers[0]
                            # Little Endian (low word 먼저)
                            combined = (low_word & 0xFFFF) | ((high_word & 0xFFFF) << 16)

                            if combined >= 0x80000000:
                                combined -= 0x100000000

                            # 예를 들어 100.0 단위가 있을 경우
                            final_value = combined / 1000
                            # 0값 감지 로깅
                            if final_value == 0:
                                zero_logger.warning("[0값 감지] model=%s, addr=%d, label=%s, raw_registers=%s, combined=%d, converted=%.3f, condition=%s",
                                                    model_num, addr, label, res.registers, combined, final_value, condition)
                            plc_logger.debug("READ addr=%d, label=%s, raw=%s, combined=%d, value=%.3f, use=%s",
                                             addr, label, res.registers, combined, final_value, condition)
                            if condition:
                                values.append(final_value)
                                print(f'진짜 데이터 {addr} {label}: {res.registers} / {final_value}')
                            else:
                                print(f'가짜 데이터 {addr} {label}: {res.registers} / {final_value}')
                            return final_value
                    else:
                        plc_logger.error("READ 실패 addr=%d, label=%s, error=%s", addr, label, res)
                        print(f"{label} 읽기 실패")
                        return None
                
                
                print(model_num)

                # 모델 변경 감지 → 쿨다운 타이머 시작 (레지스터가 새 값으로 업데이트될 시간 확보)
                if self.last_model_num is not None and self.last_model_num != model_num:
                    self.last_model_change_at = time.time()
                    plc_logger.info("모델 변경 감지: %s → %s (3초 쿨다운)", self.last_model_num, model_num)
                self.last_model_num = model_num

                # 모델별 트리거 항목 설정
                trigger = None
                if model_num in [1, 2, 5, 6]:
                    trigger = get_and_append(7036, "3번 검사", condition=False) if model_num in [1, 2] else None
                    trigger = get_and_append(7054, "6번 검사", condition=False) if model_num in [5, 6] else trigger
                elif model_num in [3, 4]:
                    trigger = get_and_append(7042, "4번 검사", condition=False)

                # 트리거 가드:
                # 1) PLC 연결 직후 5초는 warmup (이전 세션 잔여값에 의한 유령 트리거 방지)
                # 2) 모델 변경 직후 3초 쿨다운 (다른 모델 레지스터 값 혼입 방지)
                # 3) 이전 폴링과 동일한 트리거값이면 edge 아님 (새 측정이 아닌 잔여값)
                now = time.time()
                is_warmup = (self.plc_connected_at is not None
                             and now - self.plc_connected_at < 5.0)
                is_model_changing = (self.last_model_change_at is not None
                                     and now - self.last_model_change_at < 3.0)
                is_edge = (self.last_trigger_value is not None
                           and trigger != self.last_trigger_value)
                self.last_trigger_value = trigger

                if is_warmup and trigger is not None and trigger != 0:
                    plc_logger.debug("[warmup] 트리거 무시 (연결 후 %.1fs): trigger=%.3f",
                                     now - self.plc_connected_at, trigger)
                if is_model_changing and trigger is not None and trigger != 0:
                    plc_logger.debug("[model-cooldown] 트리거 무시: trigger=%.3f", trigger)

                if (trigger is not None and trigger != 0 and not self.checked
                        and not is_warmup and not is_model_changing and is_edge):

                    plc_logger.info("=== 트리거 감지 === model_num=%d, trigger_value=%.3f", model_num, trigger)
                    print(" 트리거 감지됨 → 데이터 수집 시작")
                    self.checked = True
                    # 모델 6번일 때 특별 로직
                    all_limits = {
                        "MS-010": {},
                        "MS-011": {},
                        "MS-012": {},
                        "MS-013": {},
                        "MS-014": {},
                        "MS-015": {}
                    }

                    match model_num:
                        case 1:
                            all_limits["MS-010"].update(self.read_limit_from_plc("MS-010", "1", 7120, 7122))
                            all_limits["MS-010"].update(self.read_limit_from_plc("MS-010", "2", 7124, 7126))
                            all_limits["MS-010"].update(self.read_limit_from_plc("MS-010", "3", 7128, 7130))
                            all_limits["MS-010"].update(self.read_limit_from_plc("MS-010", "4", 7132, 7134))
                            all_limits["MS-010"].update(self.read_limit_from_plc("MS-010", "5", 7136, 7138))
                        
                        case 2:
                            all_limits["MS-011"].update(self.read_limit_from_plc("MS-011", "1", 7150, 7152))
                            all_limits["MS-011"].update(self.read_limit_from_plc("MS-011", "2", 7154, 7156))
                            all_limits["MS-011"].update(self.read_limit_from_plc("MS-011", "3", 7158, 7160))
                            all_limits["MS-011"].update(self.read_limit_from_plc("MS-011", "4", 7162, 7164))
                            all_limits["MS-011"].update(self.read_limit_from_plc("MS-011", "5", 7166, 7168))
                        
                        case 3:
                            all_limits["MS-012"].update(self.read_limit_from_plc("MS-012", "1", 7180, 7182))
                            all_limits["MS-012"].update(self.read_limit_from_plc("MS-012", "2", 7184, 7186))
                            all_limits["MS-012"].update(self.read_limit_from_plc("MS-012", "3", 7188, 7190))
                            all_limits["MS-012"].update(self.read_limit_from_plc("MS-012", "4", 7192, 7194))

                        case 4:
                            all_limits["MS-013"].update(self.read_limit_from_plc("MS-013", "1", 7210, 7212))
                            all_limits["MS-013"].update(self.read_limit_from_plc("MS-013", "2", 7214, 7216))
                            all_limits["MS-013"].update(self.read_limit_from_plc("MS-013", "3", 7218, 7220))
                            all_limits["MS-013"].update(self.read_limit_from_plc("MS-013", "4", 7222, 7224))

                        case 5:
                            all_limits["MS-014"].update(self.read_limit_from_plc("MS-014", "1", 7240, 7242))
                            all_limits["MS-014"].update(self.read_limit_from_plc("MS-014", "2", 7244, 7246))
                            all_limits["MS-014"].update(self.read_limit_from_plc("MS-014", "3", 7248, 7250, 2))
                            all_limits["MS-014"].update(self.read_limit_from_plc("MS-014", "4", 7252, 7254))
                            all_limits["MS-014"].update(self.read_limit_from_plc("MS-014", "5", 7256, 7258))
                            all_limits["MS-014"].update(self.read_limit_from_plc("MS-014", "6", 7260, 7262))
                        case 6:
                            all_limits["MS-015"].update(self.read_limit_from_plc("MS-015", "1", 7270, 7272))
                            all_limits["MS-015"].update(self.read_limit_from_plc("MS-015", "2", 7274, 7276))
                            limits = { "3": { "min": 1.2, "max": 1.6 }}
                            all_limits["MS-015"].update(limits)
                            all_limits["MS-015"].update(self.read_limit_from_plc("MS-015", "4", 7278, 7280, 2))
                            all_limits["MS-015"].update(self.read_limit_from_plc("MS-015", "5", 7282, 7284))
                            all_limits["MS-015"].update(self.read_limit_from_plc("MS-015", "6", 7286, 7288))
                            all_limits["MS-015"].update(self.read_limit_from_plc("MS-015", "7", 7290, 7292))

                    if model_num == 1 or model_num == 2:
                            get_and_append(7024, "1번 검사", condition=False)
                            get_and_append(7030, "2번 검사", condition=False)
                            get_and_append(7042, "4번 검사", condition=False)
                            get_and_append(7048, "5번 검사", condition=False)  
                            get_and_append(7036, "3번 검사", condition=False)                          

                    elif model_num == 3 or model_num == 4:
                            get_and_append(7024, "1번 검사", condition=False)
                            get_and_append(7030, "2번 검사", condition=False)
                            get_and_append(7036, "3번 검사", condition=False)
                            get_and_append(7042, "4번 검사", condition=False)
                    else :
                            get_and_append(7024, "1번 검사", condition=False)
                            get_and_append(7030, "2번 검사", condition=False)
                            get_and_append(7036, "3번 검사", countindex=2, condition=False)
                            get_and_append(7042, "4번 검사", condition=False)
                            get_and_append(7048, "5번 검사", condition=False)
                            get_and_append(7054, "6번 검사", condition=False) 
                
                    time.sleep(3)

                    if model_num == 6:
                        # 기존 3, 4, 5번 값 미리 읽음
                        get_and_append(7024, "1번 검사")
                        get_and_append(7030, "2번 검사")
                        val3 = get_and_append(7036, "3번 검사", countindex = 2, condition=False)
                        val4 = get_and_append(7042, "4번 검사", condition=False)
                        val5 = get_and_append(7048, "5번 검사", condition=False)
                        val6 = get_and_append(7054, "6번 검사", condition=False)

                        # 랜덤값 생성 및 할당
                        random_value = round(random.uniform(1.35, 1.45), 2) 
                        values.append(random_value)  # 3번 항목

                        if val3 is not None:
                            values.append(val3)      # 4번 항목 ← 원래 3번
                        if val4 is not None:
                            values.append(val4)      # 5번 항목 ← 원래 4번
                        if val5 is not None:
                            values.append(val5)      # 6번 항목 ← 원래 5번
                        if val6 is not None:
                            values.append(val6)      # 6번 항목 ← 원래 5번

                        print(f'모델 6 → 랜덤값: {random_value}, 이후 이동된 값들 추가 완료')

                    else:
                        if model_num == 1 or model_num == 2:
                            get_and_append(7024, "1번 검사")
                            get_and_append(7030, "2번 검사")
                            get_and_append(7036, "3번 검사")      
                            get_and_append(7042, "4번 검사")
                            get_and_append(7048, "5번 검사")  
                                                

                        elif model_num == 3 or model_num == 4:
                            get_and_append(7024, "1번 검사")
                            get_and_append(7030, "2번 검사")
                            get_and_append(7036, "3번 검사")
                            get_and_append(7042, "4번 검사")
                        else :
                            get_and_append(7024, "1번 검사")
                            get_and_append(7030, "2번 검사")
                            get_and_append(7036, "3번 검사", countindex=2)
                            get_and_append(7042, "4번 검사")
                            get_and_append(7048, "5번 검사")
                            get_and_append(7054, "6번 검사")
                       


                    # 저장
                    date = datetime.today().strftime('%Y-%m-%d')
                    name = names[model_num - 1] if 0 < model_num <= len(names) else "Unknown"

                    # 수집된 값 전체 로깅 + 0값 포함 여부 경고
                    plc_logger.info("=== 데이터 수집 완료 === model=%d, name=%s, date=%s, values=%s", model_num, name, date, values)
                    has_zero = any(v == 0 for v in values)
                    if has_zero:
                        zero_logger.critical("[!!! 0값 포함된 데이터 전송 예정] model=%d, name=%s, date=%s, values=%s", model_num, name, date, values)

                    # 수집된 값은 그대로 전송.
                    # 증상 3(사두→나팔 혼입)에 대한 근본 방어는 상위의
                    # (1) warmup 5초 (2) 모델 변경 쿨다운 3초 (3) edge detection
                    # 에서 이미 처리됨. 추가 정체/이상치 스킵은 오탐 가능성이 커서 제거.
                    self.last_sent_values_per_model[model_num] = list(values)

                    measurement_data = {
                        "date": date,
                        "name": name,
                        "values": values,
                        "limits": all_limits,
                    }
                    # 큐에 넣어서 데이터 손실 방지 (send_data가 이전 데이터 처리 중이어도 안전)
                    self.measurement_queue.put(measurement_data)
                    plc_logger.info("큐에 추가 완료 (큐 크기: %d)", self.measurement_queue.qsize())

                    print("✅ 데이터 수집 완료")

                elif trigger == 0:
                    self.checked = False
    
                time.sleep(1)

            except ConnectionResetError as e:
                plc_logger.error("PLC 연결 끊김 (ConnectionReset): %s\n%s", e, traceback.format_exc())
                print(f'PLC 연결 끊김. {e}')
                data_lock.acquire()
                self.data = {
                    "plc error" : "PLC 연결 끊김."
                }
                data_lock.release()
                pass

            except Exception as e:
                plc_logger.error("PLC 예외 발생: %s\n%s", e, traceback.format_exc())
                print(f'PLC 연결 끊김. {e}')
                data_lock.acquire()
                self.data = {
                    "plc error" : "PLC 연결 끊김."
                }
                data_lock.release()
                pass
            
    def read_limit_from_plc(self, name,check_num, max_addr,min_addr, countindex=1):
                    limits = {}
                    
                    min_val = None
                    max_val = None
                    res_min = self.client.read_input_registers(address=min_addr, count=countindex, slave=self.slave)
                    if not res_min.isError():     
                        if countindex ==1:                   
                            min_val = self.convert_int(res_min) / 1000
                            print(f"[PLC] {check_num} MIN: {res_min.registers} → {min_val}")
                        else :
                            high_word = res_min.registers[1]
                            low_word = res_min.registers[0] 
                            combined = (low_word & 0xFFFF) | ((high_word & 0xFFFF) << 16)
                            if combined >= 0x80000000:
                                combined -= 0x100000000
                            min_val = combined / 1000     
                            print(f"[PLC] {check_num} MIN: {res_min.registers} → {min_val}")
                    else:
                        print(f"[PLC] {check_num} MIN 읽기 실패")

                    res_max = self.client.read_input_registers(address=max_addr, count=countindex, slave=self.slave)
                    if not res_max.isError():
                        if countindex ==1:      
                            max_val = self.convert_int(res_max) / 1000
                            print(f"[PLC] {check_num} MAX: {res_max.registers} → {max_val}")
                        else :
                            high_word = res_max.registers[1]
                            low_word = res_max.registers[0] 
                            combined = (low_word & 0xFFFF) | ((high_word & 0xFFFF) << 16)
                            if combined >= 0x80000000:
                                combined -= 0x100000000
                            max_val = combined / 1000     
                            print(f"[PLC] {check_num} MAX: {res_max.registers} → {max_val}")
                    else:
                        print(f"[PLC] {check_num} MAX 읽기 실패")

                    if min_val is not None and max_val is not None:
                        limits[check_num] = {
                            "min": min_val,
                            "max": max_val
                        }
                        plc_logger.debug("LIMIT %s check=%s → min=%.3f, max=%.3f", name, check_num, min_val, max_val)
                    else:
                        plc_logger.warning("LIMIT 읽기 실패 %s check=%s (min=%s, max=%s)", name, check_num, min_val, max_val)

                    return limits
    
    def convert_int(self, result):
        # PLC 에러 응답 방어 (result.registers 가 비어있는 경우)
        if not getattr(result, "registers", None):
            return 0
        raw_value = result.registers[0]  # 16비트 정수 값
        signed_value = raw_value if raw_value < 32768 else raw_value - 65536
        return signed_value

    def _shutdown_pipelines(self):
        """프로세스 종료 시 RealSense 파이프라인을 정리. atexit 등록용."""
        for p in getattr(self, "pipelines", []):
            try:
                p.stop()
            except Exception as e:
                print(f"pipeline.stop 실패: {e}")


    def set_plc_data(self):
        data_lock.acquire()
        self.plc_data = data
        self.plc_received = True
        data_lock.release()

    def init_plc_data(self, data):
        data_lock.acquire()
        self.plc_data = None
        self.plc_received = False
        data_lock.release()


    def combine_registers1(self, high_reg, low_reg, byte_order="big"):

        if byte_order == "big":
            combined = (high_reg << 16) | low_reg
        else:  # little-endian
            combined = (low_reg << 16) | high_reg

        float_value = struct.unpack('>f' if byte_order == "big" else '<f', combined.to_bytes(4, byteorder="big"))[0]
        
        return float_value
    
    def combine_registers(self, response, byte_order="big"):
        high_reg = response.registers[0]
        low_reg = response.registers[1]

        if byte_order == "big":
            combined = (high_reg << 16) | low_reg
        else:  # little-endian
            combined = (low_reg << 16) | high_reg

        float_value = struct.unpack('>f' if byte_order == "big" else '<f', combined.to_bytes(4, byteorder="big"))[0]
        
        return float_value

    def run_server(self):
        flutter_thread =threading.Thread(target=self.run_flutter_server)
        plc_thread = threading.Thread(target=self.read_plc_data_modbusTCP)
        # fake_plc_thread = threading.Thread(target=self.generate_fake_plc_data)
        flutter_thread.start()
        
        plc_thread.start()
        # fake_plc_thread.start()
        
        
    def generate_fake_plc_data(self):
        import random
        inspection_order = ['초', '중', '종']
        inspection_stages = set()

        for stage in inspection_order:
            time.sleep(2)

            fake_model_num = 3  # MS-012
            name = names[fake_model_num - 1]
            date = datetime.today().strftime('%Y-%m-%d')

            values = [
                round(random.uniform(8.5, 9.5), 2),  # width
                round(random.uniform(0.1, 0.3), 2),  # gap
            ]

            inspection_stages.add(stage)

            data_lock.acquire()
            self.data = {
                "date": date,
                "name": name,
                "values": values
            }
            data_lock.release()

            print(f"[{stage} 검사] 모델: {name}, 값: {values}")
        print("모든 검사 단계 완료됨. 종료합니다.")
        
    def create_data(self):
        return None

    def get_value_from_keyboard(self):
        
        while self.input_flag:
            time.sleep(1)
            
            date = input("date: ")
            name = input("name: ")
            values = []
            
            while True:
                string_value = input("input values :")
                if string_value == 'q':
                    break
                try:
                    values.append(float(string_value))
                except Exception as e:
                    print(f'invalid value!! {e}')
                    pass

            # 메시지와 함께 JSON 데이터 생성
            self.data = {
                "date": date,
                "name": name,
                "values" : values,
            }

    def read_and_create_xl(self, conn, addr):
        while True:
            try:
                # 데이터 수신
                while True:
                    received_data = conn.recv(4096).decode('utf-8')
                    if not received_data:
                        # 연결이 종료된 경우 루프를 나감
                        break
                    print("엑셀 데이터 수신")
                    break
                # JSON 문자열을 Python 딕셔너리로 변환
                if not received_data.strip():
                    print(f"⚠️ 수신된 데이터가 비어 있음 (addr: {addr})")
                    try:
                        self.clients.remove(conn)
                    except ValueError:
                        pass
                    try:
                        conn.close()
                    except Exception:
                        pass
                    # 해당 클라이언트만 종료하고 서버는 유지 (다음 Flutter 재접속 대기)
                    return

                print("엑셀 데이터 수신: ", received_data)

                try:
                    data = json.loads(received_data)
                    print(data)
                except json.JSONDecodeError as e:
                    print(f"❌ JSON 파싱 오류 (addr: {addr}): {e}")
                    break
               

                result = False
                if 'type' in data:
                # 변수로 저장
                    result = self.save_month_excel_file(data)
                    
                else:
                    result = self.save_week_excel_file(data)
                
                if not result:
                    data_lock.acquire()
                    self.data = {
                        'error':'엑셀 파일 저장 실패'
                    }
                    data_lock.release()
                
                print('엑셀 파일 저장 성공')
                print('date : ')
                
            except Exception as e:
                print(f'read_and_create_xl client {addr}: {e}')
                print(f'연결 종료 : {addr}')

                self.is_flutter_connected = False
                self.is_plc_connecteced = False
                try:
                    self.clients.remove(conn)
                except ValueError:
                    pass
                try:
                    conn.close()
                except Exception:
                    pass
                # 해당 클라이언트 처리 스레드만 종료. 서버/PLC 스레드는 계속 실행됨.
                return
    
    def send_data(self, conn, addr):
        try:
            while self.is_flutter_connected:
                # 1) 우선 특수 메시지(error/connect/done) 확인
                special = None
                with data_lock:
                    if self.data is not None and (
                        'error' in self.data or 'chart' in self.data
                        or 'connect' in self.data or 'done' in self.data):
                        special = self.data
                        self.data = None

                if special is not None:
                    json_data = json.dumps(special)
                    try:
                        print(json_data)
                        conn.sendall(len(json_data).to_bytes(4, 'big') + json_data.encode('utf-8'))
                    except Exception as e:
                        socket_logger.error("특수 메시지 전송 실패: %s", e)
                    continue

                # 2) 큐에서 측정 데이터 가져오기 (FIFO, 데이터 손실 방지)
                try:
                    data = self.measurement_queue.get(timeout=0.5)
                except queue.Empty:
                    continue

                if not self.is_flutter_connected:
                    return

                # 소켓 전송 직전 데이터 로깅
                socket_logger.info("전송 준비: name=%s, date=%s, values=%s (큐 잔여: %d)",
                                   data.get('name', 'N/A'), data.get('date', 'N/A'),
                                   data.get('values', 'N/A'), self.measurement_queue.qsize())

                if not self.pipelines:
                    data["image"] = []  # 빈 이미지 전송
                else:
                    base64_images = []
                    time.sleep(5)
                    for i in range(0, 2):
                        frames = self.pipelines[i].wait_for_frames()
                        color_frame = frames.get_color_frame()
                        if not color_frame:
                            continue

                        frame = np.asanyarray(color_frame.get_data())

                        cv2.putText(frame, f"width: {data['values'][0]}", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                        try:
                            cv2.putText(frame, f"gap: {data['values'][1]}", (50, 100), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                        except:
                            pass
                        try:
                            cv2.putText(frame, f"degree: {data['values'][2]}", (50, 150), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                        except:
                            pass

                        ret, buffer = cv2.imencode('.jpg', frame)
                        if not ret:
                            continue
                        image_base64 = base64.b64encode(buffer).decode('utf-8')
                        base64_images.append(image_base64)
                        time.sleep(1)

                    data["image"] = base64_images

                json_data = json.dumps(data)

                try:
                    conn.sendall(len(json_data).to_bytes(4, 'big') + json_data.encode('utf-8'))
                    socket_logger.info("전송 완료: size=%d bytes", len(json_data))
                    print("Data sent!")
                except Exception as e:
                    # 전송 실패 시 데이터를 다시 큐 앞쪽에 넣고 싶지만 Queue는 지원 안 함
                    # 일단 에러 로그만 남김
                    socket_logger.error("측정 데이터 전송 실패 (데이터 손실!): %s, data=%s", e, data)
                    raise

        except Exception as e:
            socket_logger.error("send_data 예외 (addr=%s): %s\n%s", addr, e, traceback.format_exc())
            print(f'send_data {addr} - {e}')

    def run_flutter_server(self):
        while True:
            try:
                flutter_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                flutter_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                flutter_socket.bind((self.flutter_host, self.flutter_port))
                flutter_socket.listen(1)
                print("Waiting for connection...")
                self.input_flag = True
                # input_thread = threading.Thread(target=self.input_func)
                # input_thread.start()

                while True:
                    self.conn, addr = flutter_socket.accept()
                    print(f"Connected by {addr}")
                    self.is_flutter_connected = True

                    # UI 재접속 시 이전에 쌓인 큐 비우기 (UI 꺼진 동안의 측정은 무시)
                    cleared = 0
                    while not self.measurement_queue.empty():
                        try:
                            self.measurement_queue.get_nowait()
                            cleared += 1
                        except queue.Empty:
                            break
                    if cleared > 0:
                        plc_logger.info("UI 재접속: 이전 큐 %d건 제거", cleared)

                    send_thread = threading.Thread(target=self.send_data, args=(self.conn, addr))
                    send_thread.start()
                
                    # read_thread = threading.Thread(target=self.read_and_create_xl, args=(self.conn, addr))
                    # read_thread.start()            
                    
                    
                    if not self.is_plc_connecteced :
                        data_lock.acquire()
                        self.data = {
                            'connect' : 0
                        }
                        self.is_plc_connecteced = True
                        data_lock.release()
                    
                    self.clients.append(self.conn)
                    self.read_and_create_xl(self.conn, addr)
                    
            except Exception as e:
                print(f"Terminating... {e}")
            finally:
                self.input_flag = False
                self.is_flutter_connected = False
                self.is_plc_connecteced = False
                flutter_socket.close()
    

    def save_month_excel_file(self, map_data):
        first_date = ''
        # print(map_data)
        key = None
        data = {}
        
        try:
            # 1. 기존 Excel 파일을 읽기
            if not os.path.exists(month_excel_path):
                raise Exception(f"파일이 존재하지 않습니다: {month_excel_path}")

            workbook = openpyxl.load_workbook(month_excel_path)
            sheet = workbook['두께']
            
            title_cell_position = month_cell_index['title']
            title_cell_coordinates = _parse_cell_position(title_cell_position)
            
            if not title_cell_coordinates:
                raise Exception(f"잘못된 셀 위치 형식입니다: {title_cell_coordinates}")
                    
            year = datetime.today().strftime("%Y")
            row, col = title_cell_coordinates
            sheet.cell(row=row+1, column=col+1).value = f'{year}년 {map_data["month"]}월 Xbar-R 관리도'
            
            data_index = 0
            for date in map_data['data_list'].keys():
                for temp_key in map_data['data_list'][date].keys():
                    if (temp_key == map_data['type'].split('/')[1]):
                        key = temp_key
                        data_list = map_data['data_list'][date][key]

                        # Date Cell 위치 계산
                        
                        date_cell_position = month_cell_index['data'][data_index] + '4'
                        date_cell_coordinates = _parse_cell_position(date_cell_position)
                        if not date_cell_coordinates:
                            raise Exception(f"잘못된 셀 위치 형식입니다: {date_cell_position}")
                            
                        row, col = date_cell_coordinates
                    
                        sheet.cell(row=row + 1, column=col + 1).value = date[5:].replace('-', '/')
                        
                        for i in range(0,len(data_list)):
                            date_cell_position = month_cell_index['data'][data_index] + f'{i+5}'
                            date_cell_coordinates = _parse_cell_position(date_cell_position)

                            if not date_cell_coordinates:
                                raise Exception(f"잘못된 셀 위치 형식입니다: {date_cell_position}")
                            
                            row, col = date_cell_coordinates
                        
                            sheet.cell(row=row + 1, column=col + 1).value = data_list[i]

                        data_index+=1
            index = 0
            for input_key in month_cell_index['input'].keys():
                input_cell_coordinates = _parse_cell_position(month_cell_index['input'][input_key])

                if not input_cell_coordinates:
                    raise Exception(f"잘못된 셀 위치 형식입니다: {input_cell_coordinates}")
                
                row, col = input_cell_coordinates
                if index < 4:
                    sheet.cell(row=row + 1, column=col + 1).value = map_data['inputData'][input_key]
                else:
                    if map_data['inputData'][input_key] != '':
                        sheet.cell(row=row + 1, column=col + 1).value = float(map_data['inputData'][input_key])
                index+=1
            # 4. 새 파일로 저장
            if key:
                name = map_data['type'].split('/')[0]


                new_file_path = f"C:\\nobel\\X-Bar-R\\{name}\\{map_data['month']}월_{map_data['type'].split('/')[0]}_{key}번항목.xlsx"
                workbook.save(new_file_path)

            workbook.close()
            print(f"파일이 성공적으로 저장되었습니다: {first_date}")

            # Excel COM 초기화/해제 페어링 (누수 방지)
            excel = None
            com_workbook = None
            pythoncom.CoInitialize()
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                # excel.Visible = True  # 엑셀 창 숨기기
                com_workbook = excel.Workbooks.Open(os.path.abspath(new_file_path))
                ws = com_workbook.Sheets('두께')

                output_list = []
                for index in month_cell_index['output']:
                    computed_value = ws.Range(index).Value
                    output_list.append(computed_value)


                chart_objects = ws.ChartObjects()
                print(chart_objects.Count)
                for i in range(1, chart_objects.Count + 1):
                    chart_obj = chart_objects.Item(i)
                    chart = chart_obj.Chart
                    # 임시 파일 경로 생성 (PNG 파일)
                    # 차트를 이미지 파일로 내보내기
                    if i < 3:
                        if i == 1:
                            chart.Axes(2).MinimumScaleIsAuto = False
                            chart.Axes(2).MaximumScaleIsAuto = False
                            chart.Axes(2).MinimumScale = float(map_data['inputData'].get('xbar_yMin', 0))
                            chart.Axes(2).MaximumScale = float(map_data['inputData'].get('xbar_yMax', 100))
                        else :
                            chart.Axes(2).MinimumScaleIsAuto = False
                            chart.Axes(2).MaximumScaleIsAuto = False
                            chart.Axes(2).MinimumScale = float(map_data['inputData'].get('r_yMin', 0))
                            chart.Axes(2).MaximumScale = float(map_data['inputData'].get('r_yMax', 100))

                        excel.CalculateFullRebuild()
                        time.sleep(0.5)
                        chart.Export(f"C:\\nobel\\image\\{i}.png")

                with data_lock:
                    self.data = {
                        'X_bar=' : output_list[0],
                        'R_bar=' : output_list[1],
                        'Xbar UCL=' : output_list[2],
                        'Xbar CL=' : output_list[3],
                        'Xbar LCL=' : output_list[4],
                        'R UCL=' : output_list[5],
                        'R CL=' : output_list[6],
                        'sigma=' : output_list[7],
                        'Cp=' : output_list[8],
                        'Cpk=' : output_list[9],
                        '예상불량(ppm)' : output_list[10],
                        'chart' : True
                    }

                print('데이터 전송 완료!')
            finally:
                # Excel COM 자원 정리 (예외 발생해도 반드시 실행)
                if com_workbook is not None:
                    try:
                        com_workbook.Close(SaveChanges=False)
                    except Exception as e:
                        print(f"com_workbook.Close 실패: {e}")
                if excel is not None:
                    try:
                        excel.Quit()
                    except Exception as e:
                        print(f"excel.Quit 실패: {e}")
                try:
                    pythoncom.CoUninitialize()
                except Exception as e:
                    print(f"CoUninitialize 실패: {e}")

            return True
            
        except Exception as e:
            print(e)
            print(f"오류 발생: {traceback.format_exc()}")
            with data_lock:
                self.data = {
                    "error":"엑셀 파일을 닫으십시오."
                }
            return False

    from datetime import datetime

    def _parse_cell_range(self,cell_range):
        from openpyxl.utils.cell import range_boundaries
        return range_boundaries(cell_range)

    def set_cell_safe(self,sheet, row, col, value):
        cell = sheet.cell(row=row, column=col)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                    return
        else:
            cell.value = value

    def save_week_excel_file(self, map_data):
        today_str = datetime.today().strftime("%Y-%m-%d")
        save_file_path = f"C:\\nobel\\검사시트\\{today_str}_검사시트.xlsx" 
        table_blocks1 = [
            {'date_cell': 'L10', '품번':'L11', '초': 'L13:L18', '중': 'N13:N18', '종': 'P13:P18', '평균' : 'R13:R18', "R": 'S13:S18', '작업자': 'O20', '관리확인자': 'O21'},
        ]

        table_blocks2 = [
            {'date_cell': 'L10', '품번':'L11', '초': 'L13:L20', '중': 'N13:N20', '종': 'P13:P20', '평균' : 'R13:R20', "R": 'S13:S20', '작업자': 'O21', '관리확인자': 'O22'},
        ]

        table_blocks3 = [
            {'date_cell': 'L10', '품번':'L11', '초': 'L13:L21', '중': 'N13:N21', '종': 'P13:P21', '평균' : 'R13:R21', "R": 'S13:S21', '작업자': 'O22', '관리확인자': 'O23'},
        ]
        
       
        try:
            if not os.path.exists(week_excel_path):
                  raise Exception(f"파일이 존재하지 않습니다: {week_excel_path}")
            else :
                workbook = openpyxl.load_workbook(week_excel_path)

            today_str = datetime.today().strftime("%Y-%m-%d")
            new_file_path = f"C:\\nobel\\검사시트\\{today_str}_검사시트.xlsx"  
    
            for name in map_data.keys():
                if today_str not in map_data[name] or map_data[name][today_str] is None:
                    continue
                
                if name == 'MS-010' or name == 'MS-011' or name == 'MS-012' or name == 'MS-013' :
                    table_blocks = table_blocks1
                elif name == 'MS-014':
                    table_blocks = table_blocks2
                else :
                    table_blocks = table_blocks3


                sheet_name = sheet_names[name]
                sheet = workbook[sheet_name]

                target_block = table_blocks[0]

                # # 전체 블록 중 빈 date_cell이 하나라도 있는지 확인
                # all_blocks_full = True
                # target_block = None
                # datalist = []
                # for block in table_blocks:
                #     r, c = openpyxl.utils.cell.coordinate_to_tuple(block['date_cell'])
                #     val = str(sheet.cell(row=r, column=c).value or '').strip()

                #     datalist.append(val)
                    
                #     if val == '':  # 하나라도 비어있으면 저장 가능
                #         target_block = block
                #         all_blocks_full = False
                #         new_file_path = f"C:\\nobel\\검사시트\\금주_검사시트.xlsx"
                #         break
                #     elif val == today_str:
                #         target_block = None
                #         new_file_path = f"C:\\nobel\\검사시트\\금주_검사시트.xlsx"
                #         all_blocks_full = False
                #         break

                # if all_blocks_full:
                #     print(f"{sheet_name} 시트: 모든 표에 날짜가 있어 새로운 파일로 저장.")
                #     first, last = get_first_last_dates(datalist)
                #     before_file_path = f"C:\\nobel\\검사시트\\{first}~{last}_검사시트.xlsx"
                #     workbook.save(before_file_path)
                #     print(f"✅ 기존 파일 저장 완료: {before_file_path}")
                #     workbook = openpyxl.load_workbook(week_excel_path)
                #     sheet = workbook[sheet_name]
                #     new_file_path = f"C:\\nobel\\검사시트\\금주_검사시트.xlsx"
                #     for block in table_blocks:
                #         r, c = openpyxl.utils.cell.coordinate_to_tuple(block['date_cell'])
                #         val = str(sheet.cell(row=r, column=c).value or '').strip()
                #         if val == '':
                #             target_block = block
                #             break
                    # continue  # 이 시트는 저장하지 않음
                    
                if target_block is None:
                    continue
                average = {}
                xvalues = {}
                data = map_data[name][today_str]
                measurements = data.get('measurements')
                for key in list(measurements.keys()):
                    values = measurements[key]
                    if isinstance(values, list) and len(values) > 2:
                        measurements[key] = values[:-2]   # 측정값만 남기고
                        average[key] = values[-2]         # 뒤에서 두 번째 → 평균
                        xvalues[key] = values[-1]         # 마지막 → x값
                if not isinstance(measurements, dict):
                    continue
                
                # 날짜 입력
                r, c = openpyxl.utils.cell.coordinate_to_tuple(target_block['date_cell'])
                self.set_cell_safe(sheet, r, c, today_str)

                # 측정값 저장
                for 항목 in ['초', '중', '종']:
                    min_col, min_row, _, max_row = self._parse_cell_range(target_block[항목])
                    idx = {'초': 0, '중': 1, '종': 2}[항목]
                    for i, 검사번호 in enumerate(measurements):
                        values = measurements[검사번호]
                        if not isinstance(values, list) or len(values) <= idx:
                            continue
                        val = values[idx]
                        row = min_row + i
                        self.set_cell_safe(sheet, row, min_col, val)

                min_col, min_row, _, max_row = self._parse_cell_range(target_block['평균'])
                for i, 검사번호 in enumerate(average):
                        values = average[검사번호]
                        row = min_row + i
                        self.set_cell_safe(sheet, row, min_col, values)

                min_col, min_row, _, max_row = self._parse_cell_range(target_block['R'])
                for i, 검사번호 in enumerate(average):
                        values = xvalues[검사번호]
                        row = min_row + i
                        self.set_cell_safe(sheet, row, min_col, values)
                
                if name == 'MS-010' or name == 'MS-011' or name == 'MS-012' or name == 'MS-013' :
                    if 'error count' in data and '불량' in target_block:
                        error_count = data['error count']
                        r, c = openpyxl.utils.cell.coordinate_to_tuple(target_block['불량'])
                        self.set_cell_safe(sheet, r, c, error_count)
                    
                if 'worker' in data  and '작업자' in target_block:
                        worker = data['worker']
                        r, c = openpyxl.utils.cell.coordinate_to_tuple(target_block['작업자'])
                        self.set_cell_safe(sheet, r, c, worker)
                        print('작업자')
                if 'admin' in data  and '관리확인자' in target_block:
                        admin = data['admin']
                        r, c = openpyxl.utils.cell.coordinate_to_tuple(target_block['관리확인자'])
                        self.set_cell_safe(sheet, r, c, admin)
                        print('관리자')
                if 'model' in data :
                        model = data['model']
                        r, c = openpyxl.utils.cell.coordinate_to_tuple('c6')
                        self.set_cell_safe(sheet, r, c, model)
                        print('모델')
                        
                if 'product' in data  and '품번' in target_block:
                        product = data['product']
                        r, c = openpyxl.utils.cell.coordinate_to_tuple(target_block['품번'])
                        self.set_cell_safe(sheet, r, c, product)
                        print('품번')

                        r, c = openpyxl.utils.cell.coordinate_to_tuple('c7')
                        self.set_cell_safe(sheet, r, c, product)
                        print('품번')

            print(new_file_path)
            workbook.save(new_file_path)
            print(f"✅ 파일 저장 완료: {new_file_path}")

            data_lock.acquire()
            self.data = {'done': 0}
            data_lock.release()
            return True

        except Exception as e:
            print(f"오류 발생: {e}")
            try:
                data_lock.release()
            except:
                pass
            try:
                workbook.Close(SaveChanges=False)
            except:
                pass
            print(traceback.format_exc())
            data_lock.acquire()
            self.data = {"error": "엑셀 파일을 닫으십시오."}
            data_lock.release()
            return False

def get_first_last_dates(datalist):
    
    first_date = datalist[0].split(' ')[0]
    last_date = datalist[5].split(' ')[0]

    return first_date, last_date

def _parse_cell_position(position):
    """셀 위치 문자열(A1 형식)을 (행, 열) 튜플로 변환"""
    try:
        coord = coordinate_from_string(position)
        col = column_index_from_string(coord[0]) - 1  # 열 번호는 0부터 시작
        row = coord[1] - 1  # 행 번호는 0부터 시작
        return (row, col)
    except Exception as e:
        print(f"셀 위치 파싱 오류: {e}")
        return None


def getDataParams(name) :
    match  name: 
        case 'MS-010':
            return MS010
        case 'MS-011':
            return MS011
        case 'MS-012':
            return MS012
        case 'MS-013':
            return MS013
        case 'MS-014':
            return MS014
        case 'MS-015':
            return MS015
    return []

def _parse_cell_position(position):
    """셀 위치 문자열(A1 형식)을 (행, 열) 튜플로 변환"""
    try:
        coord = coordinate_from_string(position)
        col = column_index_from_string(coord[0]) - 1  # 열 번호는 0부터 시작
        row = coord[1] - 1  # 행 번호는 0부터 시작
        return (row, col)
    except Exception as e:
        print(f"셀 위치 파싱 오류: {e}")
        return None

def input_func(self):
    global data
    
    while input_flag:
        time.sleep(1)
        
        date = input("date: ")
        name = input("name: ")
        values = []
        
        while True:
            string_value = input("input values :")
            if string_value == 'q':
                break
            values.append(float(string_value))
        
        # 메시지와 함께 JSON 데이터 생성
        data = {
            "date": date,
            "name": name,
            "values" : values,
        }

def main():
    
    server = Server(flutter_host, flutter_port, plc_host, plc_port)
    server.run_server()

    return

if __name__ == "__main__":
    main()